""" Этот код представляет собой пример использования Python для взаимодействия с YouTube Data API
и Google Drive API. Основная цель программы заключается в поиске видео на YouTube по заданному 
ключевому слову, сборе информации о найденных видео и сохранении этих данных в CSV-файл.
Кроме того, программа также может загружать данные в Google Диск при наличии соответствующих 
прав доступа.

Импортированные модули:
- openpyxl: Модуль для работы с файлами Excel (.xlsx). 
- requests: Модуль для для выполнения HTTP-запросов.
- datetime: Модуль предоставляет классы для работы с датами и временем.
- json: Модуль для преобразования между форматом JSON (JavaScript Object Notation) и объектами Python.
- logging: Модуль для логирования. 
"""


import openpyxl
from openpyxl.styles import Font, Alignment
import requests
from datetime import datetime
import json
import logging


# Настройка логгирования. Данные хранятся в файле youtube.log с уровнем логирования INFO. 
logging.basicConfig(
    filename = 'bitrix24.log',  
    level = logging.INFO,  
    format = '%(asctime)s - %(levelname)s - %(message)s', 
)


# Настройки констант для Bitrix24 API 
# (WEBHOOK - должен предоставлять доступ к определенным модулям)
BITRIX_WEBHOOK_URL = ''   


def get_candidate_data(candidate_id):
    """Функция get_candidate_data получает данные кандидата из системы Bitrix24
    по уникальному идентификатору (candidate_id).
    
    :param candidate_id: уникальный идентификатор кандидата.
    :return: словарь с информацией о кандидате.
    """
    
    logging.info(f'start get_candidate_data - {candidate_id}')
    
    url = f"{BITRIX_WEBHOOK_URL}crm.lead.get" 
    params = {
        "id": candidate_id
    }
    response = requests.get(url, params=params)
    try:
        logging.info(f'candidate details received')
        return response.json()  # Возвращаем данные кандидата в формате JSON
    except requests.exceptions.HTTPError as e:
        logging.error("HTTP error occurred: %s", e)
        return None
    except ValueError:
        logging.error("Error decoding JSON. Response:", response.text)
        return None



def save_candidate_to_excel(candidate_data, filename='candidates.xlsx'):
    """Функция предназначена для сохранения данных о кандидате в файл 
    формата Excel (*.xlsx*).
    
    :param candidate_data: словарь с данными кандидата, который будет записан в Excel-файл.
    :param filename (необязательный параметр): имя файла, куда будут сохраняться 
    данные. По умолчанию используется значение `'candidates.xlsx'`.
    return: str о том, что данные были успешно сохранены в указанный файл."""
    
    try: 
        logging.info(f'start save_candidate_to_excel')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Кандидаты'
    
        # Стили заголовков
        header_font = Font(bold=True)
        align_center = Alignment(horizontal='center')
    
        # Заголовки столбцов
        headers = ['ID', 'Имя', 'Фамилия', 'Телефон', 'Email', 'Дата создания']
        sheet.append(headers)
        for col, value in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col)
            cell.value = value
            cell.font = header_font
            cell.alignment = align_center
    
        # Заполнение данных по столбцам
        row_num = 2
        sheet.cell(row=row_num, column=1).value = candidate_data['result']['ID']
        sheet.cell(row=row_num, column=2).value = candidate_data['result']['NAME']
        sheet.cell(row=row_num, column=3).value = candidate_data['result']['LAST_NAME']
        sheet.cell(row=row_num, column=4).value = candidate_data['result']['HAS_PHONE'] if len(candidate_data['result']['HAS_PHONE']) > 0 else ''
        sheet.cell(row=row_num, column=5).value = candidate_data['result']['HAS_EMAIL'] if len(candidate_data['result']['HAS_EMAIL']) > 0 else ''
        sheet.cell(row=row_num, column=6).value = datetime.strptime(candidate_data['result']['DATE_CREATE'], '%Y-%m-%dT%H:%M:%S%z').strftime('%d.%m.%Y')
        row_num += 1

        filename = f"candidate_{candidate_data['result']['ID']}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        workbook.save(filename)
        logging.info(f'Candidate data saved in {filename}')
        return f"Данные кандидата сохранены в {filename}"
    except Exception as e:
        logging.error("Error save: %s", e)
        return None


def upload_file_to_lead(filename, candidate_id):
    """Функция предназначена для прикрепления файла с данными кандидата 
    карточки кандидата Bitrix24.
    
    :param filename: имя файла, который необходимо прикрепить к карточки
    :param candidate_id: уникальный идентификатор кандидата.
    return: str о том, что данные были успешно прикреплены"""
    
    logging.info(f'start upload_file_to_lead')
    
    # Устанавливаем ID корневой папки (0 для корневой папки)
    folder_id = 0
    
    # Загрузка файла
    with open(filename, 'rb') as file:
        file_upload_url = f"{BITRIX_WEBHOOK_URL}disk.file.upload"
        files = {'file': file}
        params = {
            'id': folder_id,  
            'name': filename 
        }
        
        # Отправляем post-запрос для загрузки файла
        upload_response = requests.post(file_upload_url, params=params, files=files)

    if upload_response.status_code == 200:
        logging.info(f'upload_file_to_lead')(f"Файл '{filename}' успешно загружен в корневую папку (ID {folder_id})!")

        # Получаем ID загруженного файла
        uploaded_file_data = upload_response.json()
        if 'result' in uploaded_file_data:
            file_id = uploaded_file_data['result']['id']
            print(f"ID загруженного файла: {file_id}")

            # Привязка файла
            attach_file_url = f"{BITRIX_WEBHOOK_URL}crm.lead.update"
            post_data = {
                'id': candidate_id,
                'fields': {
                    'UF_CRM_1644406952347': file_id  # Используйте нужный UF_CODE
                }
            }
            attach_response = requests.post(attach_file_url, json=post_data)

            if attach_response.status_code == 200:
                logging.info(f'The file was successfully attached with ID {candidate_id}')
                return f"Файл успешно прикреплён с ID {candidate_id}."
            else:
                logging.error(f"Ошибка при прикреплении файла к лиду: {attach_response.text}")
        else:
            logging.error("Ошибка получения ID загруженного файла:", upload_response.text)
    else:
        logging.error("Ошибка загрузки файла:", upload_response.text)


def read_from_excel(filename):
    """Функция предназначена чтения файла формата Excel (*.xlsx*)
    
    :param filename: имя файла, который необходимо открыть
    return: list с данными из файла. """
    
    logging.info(f'start read_from_excel')
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        data = []
        headers = next(ws.rows, None)  # Пропустить первую строку с заголовками
        for row in ws.iter_rows(min_row=2, values_only=True):
            start_date_str = row[2]
            end_date_str = row[3]
        
            # Преобразование строк в объекты datetime
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            
            data.append({
                'name': row[0],  # Имя процесса
                'description': row[1],  # Описание процесса
                'start_date': start_date.strftime('%Y-%m-%d'),  # Дата начала
                'end_date': end_date.strftime('%Y-%m-%d'),  # Дата окончания
                'responsible_user_id': row[4]  # Ответственный пользователь
            })
        logging.info(f'File read successfully - {filename}')
        print(data)
        return data
    except Exception as e:
        logging.error("Error save: %s", e)
        


def create_smart_process(data):
    """Функция предназначена создания смарт процесса в Bitrix24
    
    :param data: данные для загрузки смарт процесса (list). """
    
    logging.info(f'start create_smart_process')
    
    try: 
        smart_process_url = f'{BITRIX_WEBHOOK_URL}crm.item.add' 
        for item in data:
            payload = {
                'documentType': 'CRM_DEAL',  # Тип документа (в данном примере сделка)
                'documentId': '1',  # ID документа
                'templateId': '1',  # Укажите ID шаблона смарт-процесса
                'data': {
                    'Name': item['name'],  # Название процесса
                    'Description': item['description'],  # Описание процесса
                    'StartDate': item['start_date'],  # Дата начала
                    'EndDate': item['end_date']  # Дата окончания
                }
            }
            print('yes')
            response = requests.post(smart_process_url, json=payload)
            if response.status_code == 200:
                logging.info(f"Смарт-процесс '{item['name']}' успешно создан!")
            else:
                logging.error(f"Ошибка при создании смарт-процесса: {response.text}")
    except Exception as e:
        logging.error("Error save: %s", e)
        
     
def main():
    # Получаем данные кандидата
    candidate_id = 4 # Укажите ID кандидата, которого хотите получить
    candidate_data = get_candidate_data(candidate_id)
    print(candidate_data)

    # Генерируем Excel-файл
    excel_file = save_candidate_to_excel(candidate_data)

    print(f'Excel-файл создан: {excel_file}')
    # try:
    # # Загрузка файла на диск Bitrix24
    #     disk_file_id = upload_file_to_lead('candidate_4_20241107_1840.xlsx', candidate_data)
    #     print(f"Файл успешно загружен на диск с ID: {disk_file_id}")
    # except Exception as e:
    #     print(f"Произошла ошибка: {e}")
    # data = read_from_excel('tets_smart.xlsx')
    # print('чтение прошло успешно')
    # print(data)
    # create_smart_process(data)
    # print('смарт процесс добавлен')


if __name__ == '__main__':
    main()